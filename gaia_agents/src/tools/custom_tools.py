"""
Specific tool implementations for GAIA benchmark agent.
Each tool has a focused, single responsibility.
"""

import os
import re

import chess
import chess.pgn
import numpy as np
import pandas as pd
import requests
import whisper
import wikipedia
from bs4 import BeautifulSoup
from langchain_community.tools import DuckDuckGoSearchRun, WikipediaQueryRun
from langchain_community.tools.wikidata.tool import WikidataAPIWrapper, WikidataQueryRun
from langchain_community.utilities import ArxivAPIWrapper, WikipediaAPIWrapper
from openpyxl import load_workbook
from smolagents import tool

# === SEARCH AND WEB TOOLS ===
additional_tools = [
    WikidataQueryRun(api_wrapper=WikidataAPIWrapper()),
    WikipediaQueryRun(api_wrapper=WikipediaAPIWrapper()),
]


@tool
def wikipedia_search(query: str, language: str = "en", sentences: int = 3) -> str:
    """
    Search Wikipedia for information on a specific topic.

    Args:
        query: The search term or topic
        language: Wikipedia language code (default: "en")
        sentences: Number of sentences to return from summary (default: 3)

    Returns:
        Wikipedia article information including title, summary, and URL
    """
    try:
        wikipedia.set_lang(language)

        # Try direct page access first
        try:
            page = wikipedia.page(query)
            summary = wikipedia.summary(query, sentences=sentences)
            return f"Title: {page.title}\nSummary: {summary}\nURL: {page.url}"
        except wikipedia.DisambiguationError as e:
            # Handle disambiguation by trying first option
            page = wikipedia.page(e.options[0])
            summary = wikipedia.summary(e.options[0], sentences=sentences)
            return f"Title: {page.title}\nSummary: {summary}\nURL: {page.url}\nNote: Disambiguation resolved to first option from: {', '.join(e.options[:3])}"
        except wikipedia.PageError:
            # Search for similar pages
            search_results = wikipedia.search(query, results=5)
            if search_results:
                page = wikipedia.page(search_results[0])
                summary = wikipedia.summary(search_results[0], sentences=sentences)
                return f"Title: {page.title}\nSummary: {summary}\nURL: {page.url}\nNote: Found via search, other results: {', '.join(search_results[1:3])}"
            else:
                return f"No Wikipedia results found for: {query}"

    except Exception as e:
        return f"Error searching Wikipedia: {str(e)}"


@tool
def web_search_duckduckgo(query: str, max_results: int = 5) -> str:
    """
    Search the web using DuckDuckGo search engine.

    Args:
        query: Search query string
        max_results: Maximum number of results to return (default: 5)

    Returns:
        Formatted search results with titles, URLs, and snippets
    """
    try:
        from duckduckgo_search import DDGS

        results = []
        with DDGS() as ddgs:
            search_results = ddgs.text(query, max_results=max_results)
            for result in search_results:
                results.append(result)

        if not results:
            return f"No search results found for: {query}"

        formatted_results = []
        for i, result in enumerate(results, 1):
            formatted_results.append(
                f"{i}. {result.get('title', 'No title')}\n"
                f"   URL: {result.get('href', 'No URL')}\n"
                f"   Snippet: {result.get('body', 'No description')[:300]}...\n"
            )

        return "\n".join(formatted_results)

    except Exception as e:
        return f"Error in web search: {str(e)}"


@tool
def fetch_webpage_content(url: str, max_length: int = 3000) -> str:
    """
    Fetch and extract text content from a webpage.

    Args:
        url: The URL to fetch
        max_length: Maximum length of content to return (default: 3000)

    Returns:
        Cleaned text content from the webpage
    """
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, "html.parser")

        # Remove unwanted elements
        for element in soup(["script", "style", "nav", "header", "footer", "aside"]):
            element.decompose()

        # Extract text
        text = soup.get_text(separator=" ", strip=True)

        # Clean up whitespace
        text = re.sub(r"\s+", " ", text)

        if len(text) > max_length:
            text = text[:max_length] + "..."

        return f"Content from {url}:\n{text}"

    except Exception as e:
        return f"Error fetching webpage {url}: {str(e)}"


@tool
def arxiv_search(query: str) -> str:
    """
    Search arXiv papers.

    Args:
        query: Search query or paper ID (e.g., "1605.08386")

    Returns:
        str: arXiv paper information
    """
    try:
        arxiv = ArxivAPIWrapper()
        docs = arxiv.run(query)
        return str(docs)
    except Exception as e:
        return f"arXiv search error: {str(e)}"


@tool
def wikipedia_search_tool(query: str) -> str:
    """
    Search Wikipedia using LangChain's WikipediaQueryRun.

    Args:
        query: Search query

    Returns:
        str: Wikipedia search results
    """
    try:
        wikipedia = WikipediaQueryRun(api_wrapper=WikipediaAPIWrapper())
        result = wikipedia.run(query)
        return result
    except Exception as e:
        return f"Wikipedia search error: {str(e)}"


@tool
def duckduckgo_search(query: str) -> str:
    """
    Search using DuckDuckGo.

    Args:
        query: Search query

    Returns:
        str: DuckDuckGo search results
    """
    try:
        search = DuckDuckGoSearchRun()
        result = search.invoke(query)
        return result
    except Exception as e:
        return f"DuckDuckGo search error: {str(e)}"


@tool
def load_csv_file(
    filepath: str,
    max_rows: int = 100,
    max_columns: int = 20,
    get_all_rows: bool = False,
) -> str:
    """
    Load and analyze a CSV file.

    Args:
        filepath: Path to the CSV file
        max_rows: Maximum number of rows to display (default: 100)
        max_columns: Maximum number of columns to display (default: 20)
        get_all_rows: If True, return all rows regardless of max_rows (default: False)

    Returns:
        CSV file content and basic statistics
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File not found at {filepath}"

        # Read CSV
        df = pd.read_csv(filepath)

        # Basic info
        total_rows, total_cols = df.shape
        info = f"CSV File: {filepath}\n"
        info += f"Shape: {total_rows} rows × {total_cols} columns\n"
        info += f"Columns: {list(df.columns)}\n\n"

        # Limit display if needed
        display_rows = total_rows if get_all_rows else min(max_rows, total_rows)
        display_cols = min(max_columns, total_cols)

        if display_cols < total_cols:
            df_display = df.iloc[:display_rows, :display_cols]
            info += f"Showing first {display_rows} rows and {display_cols} columns:\n"
        else:
            df_display = df.iloc[:display_rows]
            info += f"Showing first {display_rows} rows:\n"

        info += df_display.to_string(index=True)

        # Basic statistics for numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            info += "\n\nNumeric column statistics:\n"
            info += df[numeric_cols].describe().to_string()

        return info

    except Exception as e:
        return f"Error loading CSV file: {str(e)}"


@tool
def load_excel_file(
    filepath: str,
    sheet_name: str | None = None,
    max_rows: int = 100,
    max_columns: int = 20,
    get_all_rows: bool = False,
) -> str:
    """
    Load and analyze an Excel file.

    Args:
        filepath: Path to the Excel file
        sheet_name: Specific sheet to load (default: None for first sheet)
        max_rows: Maximum number of rows to display (default: 100)
        max_columns: Maximum number of columns to display (default: 20)
        get_all_rows: If True, return all rows regardless of max_rows (default: False)

    Returns:
        Excel file content and basic statistics
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File not found at {filepath}"

        # Load workbook to get sheet names
        wb = load_workbook(filepath, read_only=True)
        sheet_names = wb.sheetnames

        # Determine which sheet to read
        if sheet_name and sheet_name in sheet_names:
            target_sheet = sheet_name
        else:
            target_sheet = sheet_names[0]  # First sheet

        # Read Excel
        df = pd.read_excel(filepath, sheet_name=target_sheet)

        # Basic info
        total_rows, total_cols = df.shape
        info = f"Excel File: {filepath}\n"
        info += f"Available sheets: {sheet_names}\n"
        info += f"Reading sheet: '{target_sheet}'\n"
        info += f"Shape: {total_rows} rows × {total_cols} columns\n"
        info += f"Columns: {list(df.columns)}\n\n"

        # Limit display if needed
        display_rows = total_rows if get_all_rows else min(max_rows, total_rows)
        display_cols = min(max_columns, total_cols)

        if display_cols < total_cols:
            df_display = df.iloc[:display_rows, :display_cols]
            info += f"Showing first {display_rows} rows and {display_cols} columns:\n"
        else:
            df_display = df.iloc[:display_rows]
            info += f"Showing first {display_rows} rows:\n"

        info += df_display.to_string(index=True)

        # Basic statistics for numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            info += "\n\nNumeric column statistics:\n"
            info += df[numeric_cols].describe().to_string()

        return info

    except Exception as e:
        return f"Error loading Excel file: {str(e)}"


@tool
def read_text_file(filepath: str, max_length: int = 2000, encoding: str = "utf-8") -> str:
    """
    Read content from a text file.

    Args:
        filepath: Path to the text file
        max_length: Maximum length of content to return (default: 2000)
        encoding: File encoding (default: "utf-8")

    Returns:
        Text file content
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: File not found at {filepath}"

        with open(filepath, encoding=encoding) as f:
            content = f.read()

        file_info = f"Text File: {filepath}\n"
        file_info += f"File size: {len(content)} characters\n"
        file_info += f"Lines: {content.count(chr(10)) + 1}\n\n"

        if len(content) > max_length:
            file_info += f"Content (first {max_length} characters):\n"
            file_info += content[:max_length] + "..."
        else:
            file_info += "Content:\n" + content

        return file_info

    except Exception as e:
        return f"Error reading text file: {str(e)}"


# === AUDIO PROCESSING ===


@tool
def transcribe_audio_file(filepath: str, model_size: str = "base") -> str:
    """
    Transcribe audio file to text using Whisper.

    Args:
        filepath: Path to the audio file
        model_size: Whisper model size ("tiny", "base", "small", "medium", "large")

    Returns:
        Transcribed text from the audio file
    """
    try:
        if not os.path.exists(filepath):
            return f"Error: Audio file not found at {filepath}"

        # Load Whisper model
        model = whisper.load_model(model_size)

        # Transcribe audio
        result = model.transcribe(filepath)

        transcription_info = f"Audio File: {filepath}\n"
        transcription_info += f"Model used: {model_size}\n"
        transcription_info += f"Language detected: {result.get('language', 'Unknown')}\n\n"
        transcription_info += f"Transcription:\n{result['text']}"

        return transcription_info

    except Exception as e:
        return f"Error transcribing audio: {str(e)}"


# === CHESS ANALYSIS ===


@tool
def analyze_chess_position(fen_notation: str) -> str:
    """
    Analyze a chess position given in FEN notation.

    Args:
        fen_notation: Chess position in FEN (Forsyth-Edwards Notation)

    Returns:
        Analysis of the chess position including legal moves
    """
    try:
        # Create board from FEN
        board = chess.Board(fen_notation)

        analysis = "Chess Position Analysis\n"
        analysis += f"FEN: {fen_notation}\n"
        analysis += f"Turn: {'White' if board.turn else 'Black'}\n"
        analysis += f"Castling rights: {board.castling_rights}\n"

        # Check game state
        if board.is_checkmate():
            analysis += "Status: CHECKMATE\n"
        elif board.is_stalemate():
            analysis += "Status: STALEMATE\n"
        elif board.is_check():
            analysis += "Status: IN CHECK\n"
        else:
            analysis += "Status: Normal play\n"

        # List legal moves
        legal_moves = list(board.legal_moves)
        analysis += f"\nNumber of legal moves: {len(legal_moves)}\n"

        if legal_moves:
            # Categorize moves
            captures = [move for move in legal_moves if board.is_capture(move)]
            checks = []
            for move in legal_moves:
                board.push(move)
                if board.is_check():
                    checks.append(move)
                board.pop()

            analysis += f"Legal moves: {', '.join([str(move) for move in legal_moves[:10]])}"
            if len(legal_moves) > 10:
                analysis += f" ... and {len(legal_moves) - 10} more"

            if captures:
                analysis += f"\nCaptures available: {', '.join([str(move) for move in captures])}"

            if checks:
                analysis += f"\nChecking moves: {', '.join([str(move) for move in checks])}"

        return analysis

    except Exception as e:
        return f"Error analyzing chess position: {str(e)}"


# === STRING MANIPULATION ===


@tool
def reverse_string(text: str) -> str:
    """
    Reverse a string character by character.

    Args:
        text: The string to reverse

    Returns:
        Reversed string
    """
    try:
        reversed_text = text[::-1]
        return f"Original: '{text}'\nReversed: '{reversed_text}'"
    except Exception as e:
        return f"Error reversing string: {str(e)}"


@tool
def reverse_words_in_string(text: str) -> str:
    """
    Reverse the order of words in a string.

    Args:
        text: The string with words to reverse

    Returns:
        String with words in reverse order
    """
    try:
        words = text.split()
        reversed_words = " ".join(words[::-1])
        return f"Original: '{text}'\nWords reversed: '{reversed_words}'"
    except Exception as e:
        return f"Error reversing words: {str(e)}"


# === DATA ANALYSIS ===


@tool
def analyze_table_commutativity(table_data: str) -> str:
    """
    Analyze a mathematical operation table for commutativity.

    Args:
        table_data: String representation of the operation table

    Returns:
        Analysis of commutativity and any counter-examples
    """
    try:
        lines = table_data.strip().split("\n")

        # Parse table
        table = {}
        elements = []

        for line in lines:
            if "|" in line and not line.strip().startswith("|---|"):
                parts = [p.strip() for p in line.split("|") if p.strip()]
                if len(parts) > 1:
                    row_element = parts[0]
                    if row_element != "*":  # Skip header row
                        elements.append(row_element)
                        table[row_element] = parts[1:]

        # Check commutativity
        violations = set()

        for i, elem1 in enumerate(elements):
            for j, elem2 in enumerate(elements):
                if i < len(table[elem1]) and j < len(table[elem2]):
                    # Get a*b and b*a
                    val1 = table[elem1][j] if j < len(table[elem1]) else None
                    val2 = table[elem2][i] if i < len(table[elem2]) else None

                    if val1 and val2 and val1 != val2:
                        violations.add(elem1)
                        violations.add(elem2)

        result = "Commutativity analysis:\n"
        result += f"Elements: {elements}\n"

        if violations:
            sorted_violations = sorted(list(violations))
            result += "Operation is NOT commutative\n"
            result += f"Elements involved in violations: {', '.join(sorted_violations)}"
        else:
            result += "Operation appears to be commutative"

        return result

    except Exception as e:
        return f"Error analyzing table commutativity: {str(e)}"


@tool
def count_items_in_list(items_text: str, separator: str = ",") -> str:
    """
    Count items in a delimited list.

    Args:
        items_text: Text containing delimited items
        separator: Delimiter to split on (default: ",")

    Returns:
        Count and list of items
    """
    try:
        items = [item.strip() for item in items_text.split(separator) if item.strip()]
        count = len(items)

        result = f"Item count: {count}\n"
        result += f"Items: {items}"

        return result

    except Exception as e:
        return f"Error counting items: {str(e)}"


@tool
def ocr_tool(image_path: str) -> str:
    """
    Extract text from images using OCR.

    Args:
        image_path: Path to image file

    Returns:
        str: Extracted text
    """
    try:
        import pytesseract
        from PIL import Image

        image = Image.open(image_path)
        text = pytesseract.image_to_string(image)
        return f"OCR text from {image_path}:\n{text.strip()}"

    except ImportError:
        return "Error: pytesseract not installed. Install with: pip install pytesseract"
    except Exception as e:
        return f"OCR error: {str(e)}"


@tool
def image_captioning_tool(image_path: str) -> str:
    """
    Generate basic image information (placeholder for actual captioning).

    Args:
        image_path: Path to image file

    Returns:
        str: Basic image information
    """
    try:
        from PIL import Image

        image = Image.open(image_path)
        width, height = image.size
        mode = image.mode
        format_type = image.format

        caption = f"Image: {image_path}\n"
        caption += f"Dimensions: {width}x{height} pixels\n"
        caption += f"Mode: {mode}\n"
        caption += f"Format: {format_type}\n"
        caption += "Note: For detailed content description, integrate with a vision model."

        return caption

    except Exception as e:
        return f"Image info error: {str(e)}"


@tool
def visual_qa_tool(image_path: str, question: str) -> str:
    """
    Answer questions about images (placeholder for actual VQA).

    Args:
        image_path: Path to image file
        question: Question about the image

    Returns:
        str: Basic response about the image
    """
    try:
        from PIL import Image

        image = Image.open(image_path)
        width, height = image.size
        mode = image.mode

        response = f"Question: {question}\n"
        response += f"Image: {image_path} ({width}x{height}, {mode})\n"
        response += "Note: For actual visual QA, integrate with a vision-language model."

        return response

    except Exception as e:
        return f"Visual QA error: {str(e)}"


all_tools = [
    wikipedia_search,
    web_search_duckduckgo,
    fetch_webpage_content,
    arxiv_search,
    wikipedia_search_tool,
    duckduckgo_search,
    load_csv_file,
    load_excel_file,
    read_text_file,
    transcribe_audio_file,
    analyze_chess_position,
    reverse_string,
    reverse_words_in_string,
    analyze_table_commutativity,
    count_items_in_list,
    ocr_tool,
    image_captioning_tool,
    visual_qa_tool,
]
